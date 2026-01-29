"""
后台管理 API 路由
提供仪表盘统计、用户管理、表单管理等管理功能
"""
import os
import json
from datetime import datetime, timedelta
from typing import Optional, List
from functools import wraps

import io
import csv
from fastapi import APIRouter, Depends, HTTPException, Query, Header
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func, desc

from database import get_db, User, InviteCode, Order, SignatureLog, SignForm, PricingPlan, UserProfile
from user_db_manager import get_user_session, ensure_user_database


# 密码文件路径
PASSWORD_FILE = os.path.join(os.path.dirname(__file__), "admin_password.json")

# 默认账号密码（从环境变量获取，优先使用文件存储）
DEFAULT_ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
DEFAULT_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")


def _read_admin_credentials_from_file():
    if os.path.exists(PASSWORD_FILE):
        try:
            with open(PASSWORD_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def get_admin_username() -> str:
    """获取当前管理员账号（优先从文件读取）"""
    data = _read_admin_credentials_from_file()
    return data.get("username") or DEFAULT_ADMIN_USERNAME


def get_admin_password() -> str:
    """获取当前管理员密码（优先从文件读取）"""
    data = _read_admin_credentials_from_file()
    return data.get("password", DEFAULT_PASSWORD)


def save_admin_credentials(new_username: str, new_password: str) -> bool:
    """保存新账号/密码到文件"""
    try:
        with open(PASSWORD_FILE, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "username": new_username,
                    "password": new_password,
                    "updated_at": datetime.now().isoformat(),
                },
                f,
                ensure_ascii=False,
            )
        return True
    except Exception as e:
        print(f"保存账号密码失败: {e}")
        return False


router = APIRouter(
    prefix="/admin",
    tags=["管理后台"],
    responses={401: {"description": "未授权"}},
)


# ==================== 认证 ====================

def verify_admin(
    x_admin_username: str = Header(None),
    x_admin_token: str = Header(None),
    token: str = Query(None),
):
    """验证管理员身份

    - 新方式：Header 同时带上 X-Admin-Username + X-Admin-Token
    - 兼容旧方式：Header X-Admin-Token 或 Query token（仅密码）
    """
    current_password = get_admin_password()
    current_username = get_admin_username()

    # 旧兼容：只校验密码
    if x_admin_username is None:
        actual_token = x_admin_token or token
        if actual_token != current_password:
            raise HTTPException(status_code=401, detail="管理员认证失败")
        return True

    # 新方式：账号+密码
    if x_admin_username != current_username:
        raise HTTPException(status_code=401, detail="管理员认证失败")

    if x_admin_token != current_password:
        raise HTTPException(status_code=401, detail="管理员认证失败")

    return True


# ==================== 密码管理 ====================

class ChangePasswordRequest(BaseModel):
    """修改密码请求"""

    old_password: str
    new_password: str


@router.put("/password", summary="修改密码")
def change_password(req: ChangePasswordRequest, _: bool = Depends(verify_admin)):
    """修改管理员密码"""
    current_password = get_admin_password()

    if req.old_password != current_password:
        raise HTTPException(status_code=400, detail="旧密码错误")

    if len(req.new_password) < 6:
        raise HTTPException(status_code=400, detail="新密码至少6位")

    if req.new_password == req.old_password:
        raise HTTPException(status_code=400, detail="新密码不能与旧密码相同")

    if save_admin_credentials(get_admin_username(), req.new_password):
        return {"success": True, "message": "密码修改成功"}
    else:
        raise HTTPException(status_code=500, detail="密码保存失败")


class ChangeCredentialsRequest(BaseModel):
    """修改账号密码请求"""

    old_username: str
    old_password: str
    new_username: str
    new_password: str


@router.put("/credentials", summary="修改账号密码")
def change_credentials(req: ChangeCredentialsRequest, _: bool = Depends(verify_admin)):
    current_username = get_admin_username()
    current_password = get_admin_password()

    if req.old_username != current_username or req.old_password != current_password:
        raise HTTPException(status_code=400, detail="旧账号或密码错误")

    if len(req.new_username) < 3:
        raise HTTPException(status_code=400, detail="新账号至少3位")

    if len(req.new_password) < 6:
        raise HTTPException(status_code=400, detail="新密码至少6位")

    if save_admin_credentials(req.new_username, req.new_password):
        return {"success": True, "message": "账号密码修改成功"}
    else:
        raise HTTPException(status_code=500, detail="账号密码保存失败")


# ==================== 响应模型 ====================

class DashboardStats(BaseModel):
    """仪表盘统计数据"""

    total_users: int
    new_users_today: int
    total_signatures: int
    signatures_today: int
    active_forms: int
    total_form_submissions: int
    total_invites: int
    active_invites: int
    total_income: int  # 总收入（分）
    income_today: int  # 今日收入（分）


class UserItem(BaseModel):
    """用户列表项"""

    id: int
    open_id: str
    tenant_key: str
    remaining_quota: int
    total_used: int
    invite_code_used: Optional[str]
    invite_expire_at: Optional[datetime]
    total_paid: int
    created_at: datetime


class FormItem(BaseModel):
    """表单列表项"""

    id: int
    form_id: str
    name: str
    description: Optional[str]
    submit_count: int
    is_active: bool
    created_by: Optional[str]
    created_at: datetime


class InviteItem(BaseModel):
    """邀请码列表项"""

    id: int
    code: str
    max_usage: int
    used_count: int
    benefit_days: int
    expires_at: Optional[datetime]
    is_active: bool
    created_by: str
    created_at: datetime


class LogItem(BaseModel):
    """签名日志项"""

    id: int
    user_key: str
    file_name: Optional[str]
    file_token: Optional[str]
    quota_consumed: bool
    created_at: datetime


class UpdateQuotaRequest(BaseModel):
    """更新配额请求"""

    remaining_quota: int


class CreateInviteRequest(BaseModel):
    """创建邀请码请求"""

    max_usage: int = 10
    benefit_days: int = 365
    expires_in_days: int = 30


# ==================== 仪表盘 API ====================

@router.get("/dashboard", response_model=DashboardStats, summary="仪表盘统计")
def get_dashboard(db: Session = Depends(get_db), _: bool = Depends(verify_admin)):
    """获取仪表盘统计数据"""
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)

    # 用户统计
    total_users = db.query(User).count()
    new_users_today = db.query(User).filter(User.created_at >= today).count()

    # 签名统计
    total_signatures = db.query(SignatureLog).count()
    signatures_today = db.query(SignatureLog).filter(SignatureLog.created_at >= today).count()

    # 表单统计
    active_forms = db.query(SignForm).filter(SignForm.is_active == True).count()
    total_form_submissions = db.query(func.sum(SignForm.submit_count)).scalar() or 0

    # 邀请码统计
    total_invites = db.query(InviteCode).count()
    active_invites = db.query(InviteCode).filter(InviteCode.is_active == True).count()

    # 收入统计（只统计已支付的订单）
    total_income = db.query(func.sum(Order.amount)).filter(Order.status == "paid").scalar() or 0
    income_today = (
        db.query(func.sum(Order.amount))
        .filter(Order.status == "paid", Order.paid_at >= today)
        .scalar()
        or 0
    )

    return DashboardStats(
        total_users=total_users,
        new_users_today=new_users_today,
        total_signatures=total_signatures,
        signatures_today=signatures_today,
        active_forms=active_forms,
        total_form_submissions=total_form_submissions,
        total_invites=total_invites,
        active_invites=active_invites,
        total_income=total_income,
        income_today=income_today,
    )


@router.get("/dashboard/trends", summary="趋势数据")
def get_dashboard_trends(
    period: str = Query("week", description="时间周期: week(本周) 或 month(本月)"),
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """获取仪表盘趋势数据，用于图表展示

    返回按日期分组的用户数和签名次数
    """
    from sqlalchemy import cast, Date

    # 计算起始日期
    today = datetime.utcnow().date()
    if period == "month":
        start_date = today - timedelta(days=29)
    else:
        start_date = today - timedelta(days=6)

    date_range = []
    current = start_date
    while current <= today:
        date_range.append(current.isoformat())
        current += timedelta(days=1)

    user_counts = {}
    user_results = (
        db.query(cast(User.created_at, Date).label("date"), func.count(User.id).label("count"))
        .filter(User.created_at >= datetime.combine(start_date, datetime.min.time()))
        .group_by(cast(User.created_at, Date))
        .all()
    )

    for row in user_results:
        date_str = row.date.isoformat() if hasattr(row.date, "isoformat") else str(row.date)
        user_counts[date_str] = row.count

    sig_counts = {}
    sig_results = (
        db.query(
            cast(SignatureLog.created_at, Date).label("date"),
            func.count(SignatureLog.id).label("count"),
        )
        .filter(SignatureLog.created_at >= datetime.combine(start_date, datetime.min.time()))
        .group_by(cast(SignatureLog.created_at, Date))
        .all()
    )

    for row in sig_results:
        date_str = row.date.isoformat() if hasattr(row.date, "isoformat") else str(row.date)
        sig_counts[date_str] = row.count

    users_data = [user_counts.get(d, 0) for d in date_range]
    signatures_data = [sig_counts.get(d, 0) for d in date_range]

    return {"dates": date_range, "users": users_data, "signatures": signatures_data}


# ==================== 用户管理 API ====================

@router.get("/users", summary="用户列表")
def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """获取用户列表（分页）"""
    query = db.query(User)

    if search:
        query = query.filter(User.open_id.contains(search))

    total = query.count()
    users = (
        query.order_by(desc(User.created_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": u.id,
                "open_id": u.open_id,
                "tenant_key": u.tenant_key,
                "remaining_quota": u.remaining_quota,
                "total_used": u.total_used,
                "invite_code_used": u.invite_code_used,
                "invite_expire_at": u.invite_expire_at.isoformat() if u.invite_expire_at else None,
                "total_paid": u.total_paid,
                "created_at": u.created_at.isoformat(),
            }
            for u in users
        ],
    }


@router.put("/users/{user_id}/quota", summary="调整用户配额")
def update_user_quota(
    user_id: int,
    req: UpdateQuotaRequest,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """调整用户配额"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    user.remaining_quota = req.remaining_quota
    db.commit()

    return {"success": True, "remaining_quota": user.remaining_quota}


@router.delete("/users/{user_id}", summary="删除用户")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """删除用户及其关联记录"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    db.delete(user)
    db.commit()

@router.post("/users/{user_id}/reset", summary="初始化/重置用户数据")
def reset_user(
    user_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """重置用户数据（包括配额、邀请码、已用次数等）"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 1. 处理旧的邀请码计数（如果需要）
    if user.invite_code_used:
        old_invite = db.query(InviteCode).filter(InviteCode.code == user.invite_code_used).first()
        if old_invite and old_invite.used_count > 0:
            old_invite.used_count -= 1
    
    # 2. 重置用户库状态
    try:
        ensure_user_database(user.user_key)
        user_db = get_user_session(user.user_key)
        
        if "::" in user.user_key:
            open_id_val, tenant_key_val = user.user_key.split("::")
            user_profile = user_db.query(UserProfile).filter(
                UserProfile.open_id == open_id_val,
                UserProfile.tenant_key == tenant_key_val
            ).first()
            
            if user_profile:
                user_profile.remaining_quota = 20  # 恢复默认 20 次
                user_profile.total_used = 0
                user_profile.invite_code_used = None
                user_profile.invite_expire_at = None
                user_profile.current_plan_id = None  # 清除套餐
                user_profile.plan_expires_at = None
                user_profile.plan_quota_reset_at = None
                user_profile.is_unlimited = False
                user_db.commit()
        user_db.close()
    except Exception as e:
        print(f"Failed to reset user profile for {user.user_key}: {e}")
        # 这里不阻断，继续重置共享库

    # 3. 重置共享库状态
    user.remaining_quota = 20
    user.total_used = 0
    user.invite_code_used = None
    user.invite_expire_at = None
    user.total_paid = 0
    
    db.commit()
    
    return {"success": True, "message": "用户数据已初始化"}


@router.delete("/users/{user_id}/invite", summary="清理用户邀请码记录")
def clear_user_invite(
    user_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """清理用户的邀请码使用记录"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    if not user.invite_code_used:
        return {"success": True, "message": "用户未使用过邀请码", "invite_code": None}

    invite_code_str = user.invite_code_used

    user.invite_code_used = None
    user.invite_expire_at = None

    invite = db.query(InviteCode).filter(InviteCode.code == invite_code_str).first()
    if invite and invite.used_count > 0:
        invite.used_count -= 1

    db.commit()

    return {
        "success": True,
        "message": "邀请码记录已清理",
        "invite_code": invite_code_str,
        "used_count_decreased": invite.used_count if invite else None,
    }


@router.get("/users/export", summary="导出用户 CSV")
def export_users(db: Session = Depends(get_db), _: bool = Depends(verify_admin)):
    users = db.query(User).all()

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器未安装 openpyxl 库，无法导出 Excel。请联系管理员安装。")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户列表"

    headers = ["ID", "OpenID", "租户Key", "免费次数", "已使用", "邀请码", "VIP到期时间", "累计付费", "注册时间"]
    ws.append(headers)

    for u in users:
        ws.append(
            [
                u.id,
                u.open_id,
                u.tenant_key,
                u.remaining_quota,
                u.total_used,
                u.invite_code_used or "-",
                u.invite_expire_at.isoformat() if u.invite_expire_at else "无",
                u.total_paid / 100.0,
                u.created_at.isoformat(),
            ]
        )

    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = (
        f"attachment; filename=users_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    )
    return response


# ==================== 表单管理 API ====================

@router.get("/forms", summary="表单列表")
def list_forms(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """获取表单列表（分页）"""
    query = db.query(SignForm)

    if search:
        query = query.filter(SignForm.name.contains(search))

    total = query.count()
    forms = (
        query.order_by(desc(SignForm.created_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": f.id,
                "form_id": f.form_id,
                "name": f.name,
                "description": f.description,
                "submit_count": f.submit_count,
                "is_active": f.is_active,
                "created_by": f.created_by,
                "created_at": f.created_at.isoformat(),
            }
            for f in forms
        ],
    }


@router.put("/forms/{form_id}/status", summary="更新表单状态")
def update_form_status(
    form_id: str,
    is_active: bool = Query(...),
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """启用或禁用表单"""
    form = db.query(SignForm).filter(SignForm.form_id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="表单不存在")

    form.is_active = is_active
    db.commit()

    return {"success": True, "is_active": form.is_active}


@router.delete("/forms/{form_id}", summary="删除表单")
def delete_form(
    form_id: str,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """永久删除外部签名表单"""
    form = db.query(SignForm).filter(SignForm.form_id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="表单不存在")

    db.delete(form)
    db.commit()

    return {"success": True, "message": "表单已删除"}


# ==================== 邀请码管理 API ====================

@router.get("/invites", summary="邀请码列表")
def list_invites(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """获取邀请码列表（分页）"""
    query = db.query(InviteCode)

    total = query.count()
    invites = (
        query.order_by(desc(InviteCode.created_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": i.id,
                "code": i.code,
                "max_usage": i.max_usage,
                "used_count": i.used_count,
                "benefit_days": i.benefit_days,
                "expires_at": i.expires_at.isoformat() if i.expires_at else None,
                "is_active": i.is_active,
                "created_by": i.created_by,
                "created_at": i.created_at.isoformat(),
            }
            for i in invites
        ],
    }


@router.post("/invites", summary="创建邀请码")
def create_invite(req: CreateInviteRequest, db: Session = Depends(get_db), _: bool = Depends(verify_admin)):
    """创建新邀请码"""
    import uuid

    code = f"INV-{uuid.uuid4().hex[:8].upper()}"
    expires_at = datetime.utcnow() + timedelta(days=req.expires_in_days) if req.expires_in_days else None

    invite = InviteCode(
        code=code,
        max_usage=req.max_usage,
        benefit_days=req.benefit_days,
        expires_at=expires_at,
        created_by="admin",
    )
    db.add(invite)
    db.commit()
    db.refresh(invite)

    return {"success": True, "code": code, "expires_at": expires_at.isoformat() if expires_at else None}


@router.delete("/invites/{invite_id}", summary="删除邀请码")
def delete_invite(invite_id: int, db: Session = Depends(get_db), _: bool = Depends(verify_admin)):
    """删除邀请码"""
    invite = db.query(InviteCode).filter(InviteCode.id == invite_id).first()
    if not invite:
        raise HTTPException(status_code=404, detail="邀请码不存在")

    db.delete(invite)
    db.commit()
    return {"success": True, "message": "邀请码已删除"}


@router.get("/invites/export", summary="导出邀请码 CSV")
def export_invites(db: Session = Depends(get_db), _: bool = Depends(verify_admin)):
    """导出所有邀请码为 CSV"""
    invites = db.query(InviteCode).all()

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器未安装 openpyxl 库，无法导出 Excel。请联系管理员安装。")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "邀请码列表"

    headers = ["ID", "邀请码", "最大使用", "已使用", "权益天数", "领取截止日期", "状态", "创建人", "创建时间"]
    ws.append(headers)

    for i in invites:
        ws.append(
            [
                i.id,
                i.code,
                i.max_usage,
                i.used_count,
                i.benefit_days,
                i.expires_at.isoformat() if i.expires_at else "永久",
                "有效" if i.is_active else "禁用",
                i.created_by,
                i.created_at.isoformat(),
            ]
        )

    for idx, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = (
        f"attachment; filename=invites_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    )
    return response


@router.put("/invites/{invite_id}/status", summary="更新邀请码状态")
def update_invite_status(
    invite_id: int,
    is_active: bool = Query(...),
    revoke_benefits: bool = Query(False, description="禁用时是否撤销已兑换用户的权益"),
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """启用或禁用邀请码"""
    invite = db.query(InviteCode).filter(InviteCode.id == invite_id).first()
    if not invite:
        raise HTTPException(status_code=404, detail="邀请码不存在")

    invite.is_active = is_active

    revoked_count = 0

    if not is_active and revoke_benefits:
        affected_users = db.query(User).filter(User.invite_code_used == invite.code).all()
        for user in affected_users:
            # 1. 更新共享库 legacy 状态（保留原有逻辑）
            if user.invite_expire_at:
                user.invite_expire_at = None
                revoked_count += 1
            
            # 2. 更新独立用户库状态
            try:
                ensure_user_database(user.user_key)
                user_db = get_user_session(user.user_key)
                # 分解 open_id 和 tenant_key
                # user_key format: "open_id::tenant_key"
                if "::" in user.user_key:
                    open_id_val, tenant_key_val = user.user_key.split("::")
                    user_profile = user_db.query(UserProfile).filter(
                        UserProfile.open_id == open_id_val,
                        UserProfile.tenant_key == tenant_key_val
                    ).first()
                    
                    if user_profile and user_profile.invite_code_used == invite.code:
                        user_profile.invite_expire_at = None
                        # 如果需要连 invite_code_used 也清除，可以解开下行注释
                        # user_profile.invite_code_used = None 
                        user_db.commit()
                user_db.close()
            except Exception as e:
                print(f"Failed to revoke benefit for user {user.user_key}: {e}")

    db.commit()

    return {"success": True, "is_active": invite.is_active, "revoked_count": revoked_count}


# ==================== 签名日志 API ====================

@router.get("/logs", summary="签名日志")
def list_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user_key: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """获取签名日志列表（分页）"""
    query = db.query(SignatureLog)

    if user_key:
        query = query.filter(SignatureLog.user_key.contains(user_key))

    if start_date:
        try:
            start = datetime.fromisoformat(start_date)
            query = query.filter(SignatureLog.created_at >= start)
        except Exception:
            pass

    if end_date:
        try:
            end = datetime.fromisoformat(end_date)
            query = query.filter(SignatureLog.created_at <= end)
        except Exception:
            pass

    total = query.count()
    logs = (
        query.order_by(desc(SignatureLog.created_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": l.id,
                "user_key": l.user_key,
                "file_name": l.file_name,
                "file_token": l.file_token,
                "quota_consumed": l.quota_consumed,
                "created_at": l.created_at.isoformat(),
            }
            for l in logs
        ],
    }


# ==================== 订单管理 API ====================

@router.get("/orders", summary="订单列表")
def list_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    """获取订单列表（分页）"""
    query = db.query(Order)

    if status:
        query = query.filter(Order.status == status)

    total = query.count()
    orders = (
        query.order_by(desc(Order.created_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": o.id,
                "order_id": o.order_id,
                "user_key": o.user_key,
                "plan_id": o.plan_id,
                "quota_count": o.quota_count,
                "amount": o.amount,
                "status": o.status,
                "payment_method": o.payment_method,
                "created_at": o.created_at.isoformat() if o.created_at else None,
                "paid_at": o.paid_at.isoformat() if o.paid_at else None,
            }
            for o in orders
        ],
    }


# ==================== 套餐管理 API ====================

class CreatePricingPlanRequest(BaseModel):
    """创建套餐请求"""

    plan_id: str
    name: str
    quota_count: int
    price: int
    sort_order: int = 0
    description: Optional[str] = None
    billing_type: Optional[str] = None
    monthly_price: Optional[int] = None
    yearly_price: Optional[int] = None
    unlimited: Optional[bool] = None
    save_percent: Optional[int] = None


class UpdatePricingPlanRequest(BaseModel):
    """更新套餐请求"""

    name: Optional[str] = None
    quota_count: Optional[int] = None
    price: Optional[int] = None
    sort_order: Optional[int] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None
    billing_type: Optional[str] = None
    monthly_price: Optional[int] = None
    yearly_price: Optional[int] = None
    unlimited: Optional[bool] = None
    save_percent: Optional[int] = None


def _sync_pricing_derived_fields(plan: PricingPlan) -> None:
    if plan.billing_type == "monthly":
        if plan.monthly_price is None:
            plan.monthly_price = plan.price
    elif plan.billing_type == "yearly":
        if plan.yearly_price is None:
            plan.yearly_price = plan.price
    else:
        plan.billing_type = "monthly"
        if plan.monthly_price is None:
            plan.monthly_price = plan.price

    if plan.monthly_price and plan.yearly_price and plan.monthly_price > 0:
        yearly_from_monthly = plan.monthly_price * 12
        if yearly_from_monthly > 0 and plan.yearly_price and plan.yearly_price > 0:
            saved = max(0.0, 1.0 - (plan.yearly_price / float(yearly_from_monthly)))
            plan.save_percent = int(round(saved * 100))


@router.get("/pricing", summary="套餐列表")
def list_pricing_plans(
    include_inactive: bool = Query(False, description="是否包含已下架套餐"),
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    query = db.query(PricingPlan)
    if not include_inactive:
        query = query.filter(PricingPlan.is_active == True)

    plans = query.order_by(PricingPlan.sort_order).all()

    return {
        "items": [
            {
                "id": p.id,
                "plan_id": p.plan_id,
                "name": p.name,
                "quota_count": p.quota_count,
                "price": p.price,
                "billing_type": p.billing_type,
                "monthly_price": p.monthly_price,
                "yearly_price": p.yearly_price,
                "unlimited": p.unlimited,
                "save_percent": p.save_percent,
                "is_active": p.is_active,
                "sort_order": p.sort_order,
                "description": p.description,
                "created_at": p.created_at.isoformat() if p.created_at else None,
                "updated_at": p.updated_at.isoformat() if p.updated_at else None,
            }
            for p in plans
        ]
    }


@router.post("/pricing", summary="创建套餐")
def create_pricing_plan(
    req: CreatePricingPlanRequest,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    existing = db.query(PricingPlan).filter(PricingPlan.plan_id == req.plan_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="套餐ID已存在")

    if req.price < 0:
        raise HTTPException(status_code=400, detail="价格不能为负数")
    if req.quota_count is not None and req.quota_count < 0:
        raise HTTPException(status_code=400, detail="签名次数不能为负数")

    plan = PricingPlan(
        plan_id=req.plan_id,
        name=req.name,
        quota_count=req.quota_count,
        price=req.price,
        sort_order=req.sort_order,
        description=req.description,
        is_active=True,
        billing_type=req.billing_type or "monthly",
        monthly_price=req.monthly_price,
        yearly_price=req.yearly_price,
        unlimited=bool(req.unlimited) if req.unlimited is not None else False,
        save_percent=req.save_percent,
    )
    _sync_pricing_derived_fields(plan)
    db.add(plan)
    db.commit()
    db.refresh(plan)

    return {"success": True, "plan_id": plan.plan_id, "id": plan.id}


@router.put("/pricing/{plan_id}", summary="更新套餐")
def update_pricing_plan(
    plan_id: str,
    req: UpdatePricingPlanRequest,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    plan = db.query(PricingPlan).filter(PricingPlan.plan_id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="套餐不存在")

    if req.name is not None:
        plan.name = req.name
    if req.quota_count is not None:
        plan.quota_count = req.quota_count
    if req.price is not None:
        if req.price < 0:
            raise HTTPException(status_code=400, detail="价格不能为负数")
        plan.price = req.price
    if req.sort_order is not None:
        plan.sort_order = req.sort_order
    if req.description is not None:
        plan.description = req.description
    if req.is_active is not None:
        plan.is_active = req.is_active
    if req.billing_type is not None:
        if req.billing_type not in ("monthly", "yearly"):
            raise HTTPException(status_code=400, detail="billing_type 只能是 monthly 或 yearly")
        plan.billing_type = req.billing_type
    if req.monthly_price is not None:
        if req.monthly_price < 0:
            raise HTTPException(status_code=400, detail="monthly_price 不能为负数")
        plan.monthly_price = req.monthly_price
    if req.yearly_price is not None:
        if req.yearly_price < 0:
            raise HTTPException(status_code=400, detail="yearly_price 不能为负数")
        plan.yearly_price = req.yearly_price
    if req.unlimited is not None:
        plan.unlimited = req.unlimited
    if req.save_percent is not None:
        if req.save_percent < 0 or req.save_percent > 100:
            raise HTTPException(status_code=400, detail="save_percent 必须在 0-100 之间")
        plan.save_percent = req.save_percent

    if req.price is not None:
        if plan.billing_type == "monthly" and req.monthly_price is None:
            plan.monthly_price = plan.price
        if plan.billing_type == "yearly" and req.yearly_price is None:
            plan.yearly_price = plan.price

    _sync_pricing_derived_fields(plan)
    db.commit()

    return {"success": True, "plan_id": plan.plan_id}


@router.delete("/pricing/{plan_id}", summary="删除套餐")
def delete_pricing_plan(
    plan_id: str,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin),
):
    plan = db.query(PricingPlan).filter(PricingPlan.plan_id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="套餐不存在")

    db.delete(plan)
    db.commit()

    return {"success": True, "message": f"套餐 {plan_id} 已删除"}
